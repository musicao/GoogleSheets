import requests
import re
from io import BytesIO
from openpyxl import load_workbook

def download_file_from_google_drive(id):
    def get_confirm_token(response):
        for key, value in response.cookies.items():
            if key.startswith('download_warning'):
                return value

        return None

    def save_response_content(response):
        CHUNK_SIZE = 32768
        buff = BytesIO()

        for chunk in response.iter_content(CHUNK_SIZE):
            if chunk:
                buff.write(chunk)

        return buff

    def factory(buff,padding,botton):

        obj = []
        wb = load_workbook(buff)
        planilha = wb[wb.sheetnames[0]]
        for nlinha, linha in enumerate(planilha.rows):
            if nlinha <= padding or nlinha >= botton:
                continue
            processamento = {d.column: d.value if hasattr(d, 'value') else None for d in linha}

            obj.append(processamento)

        return obj

    URL = "https://docs.google.com/uc?export=download"

    session = requests.Session()

    response = session.get(URL, params={'id': id}, stream=True)
    token = get_confirm_token(response)

    if token:
        params = {'id': id, 'confirm': token}
        response = session.get(URL, params=params, stream=True)

    buff = save_response_content(response)
    return  factory(buff, padding = 2,botton=1672)




class GoogleSheets():

    @staticmethod
    def download(file_id):
        return download_file_from_google_drive(file_id)



